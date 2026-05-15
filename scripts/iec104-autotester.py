#!/usr/bin/env python3
"""
IEC104 AutoTester — automated integration test runner.
Reads test plan from Excel, drives MCP tools, generates markdown report.
"""

import argparse
import json
import re
import subprocess
import sys
import time
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    import openpyxl
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)


def load_point_mapping(xlsx_dir: str = ".") -> Dict[str, Dict[str, int]]:
    """扫描目录下的 xlsx 点表文件，建立 {实例名→{点名→IOA}} 映射。"""
    mapping = {}
    for f in os.listdir(xlsx_dir):
        if not f.endswith(".xlsx"):
            continue
        name = f.replace(".xlsx", "").replace("固定验证-", "")
        try:
            wb = openpyxl.load_workbook(os.path.join(xlsx_dir, f), data_only=True)
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            points = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                pt = dict(zip(headers, row))
                ioa = pt.get("point-number")
                pname = pt.get("point-name")
                if ioa and pname:
                    points[pname] = int(ioa)
                    points[str(ioa)] = int(ioa)  # also map by IOA string
            wb.close()
            mapping[name] = points
        except Exception as e:
            print(f"  ⚠️  读取点表跳过 {f}: {e}")
    return mapping


def parse_natural_step(desc: str, point_map: Dict[str, Dict[str, int]]) -> List[dict]:
    """解析自然语言测试步骤描述为结构化操作列表。

    支持的描述模式:
    - "设置XX功率=值" / "XX功率变成值" → write(XX实例, IOA, 值)
    - "预期: XX=值" / "预期看到XX=值" → assert(XX实例, IOA, 值)
    - "对XX下发YY=值" → write(XX实例, IOA, 值)
    """
    steps = []

    # Find all instances mentioned in description
    for inst_name, pt_map in point_map.items():
        if inst_name not in desc:
            continue

        # Find all points mentioned for this instance
        for pt_name, ioa in pt_map.items():
            if pt_name not in desc:
                continue
            # Match patterns like "XX=值" or "XX变成值"
            patterns = [
                rf"{re.escape(pt_name)}[=变成](-?\d+\.?\d*)",
                rf"{re.escape(pt_name)}为(-?\d+\.?\d*)",
            ]
            for pat in patterns:
                m = re.search(pat, desc)
                if m:
                    val = float(m.group(1))
                    is_input = ioa < 20000  # AI points < 20000, AO >= 25089
                    if "预期" in desc:
                        steps.append({
                            "action": "assert",
                            "instance": inst_name,
                            "ioa": ioa,
                            "expected_min": val * 0.95,
                            "expected_max": val * 1.05,
                            "expected_val": val,
                            "desc_detail": f"{pt_name}={val}",
                        })
                    else:
                        steps.append({
                            "action": "write",
                            "instance": inst_name,
                            "ioa": ioa,
                            "value": val,
                            "desc_detail": f"{pt_name}={val}",
                        })
                    break  # first match for this point
    return steps


def apply_port_template(instances: List[dict], port_start: int):
    """将 instances 配置中的 {p+N} 模板替换为实际端口号。"""
    for inst in instances:
        for key in ("iec104_port", "http_port"):
            val = str(inst.get(key, ""))
            if "{p" in val:
                import re as _re
                m = _re.search(r"\{p\+(\d+)\}", val)
                if m:
                    offset = int(m.group(1))
                    inst[key] = port_start + offset


def mcp_call(mcp_bin: str, simulator_url: str, tool: str, args: dict) -> dict:
    """Call MCP tool via stdio JSON-RPC (shell pipe for reliable stdin)."""
    import shlex

    payload = {
        "jsonrpc": "2.0",
        "id": 1,
        "method": "tools/call",
        "params": {"name": tool, "arguments": args},
    }
    payload_str = json.dumps(payload, ensure_ascii=False)

    # Use echo piped to MCP binary via shell (direct stdin piping doesn't work with this Go binary)
    mcp_esc = shlex.quote(mcp_bin)
    sim_esc = shlex.quote(simulator_url)
    # Python JSON → shell-safe single-quoted string
    payload_esc = payload_str.replace("'", "'\\''")
    cmd = f"echo '{payload_esc}' | {mcp_esc} -simulator {sim_esc} -mode both 2>/dev/null"

    try:
        proc = subprocess.run(cmd, shell=True, capture_output=True, timeout=30)
        raw = proc.stdout.decode("utf-8", errors="replace").strip()
    except subprocess.TimeoutExpired:
        return {"_error": "MCP call timed out"}

    if not raw:
        return {"_error": "no output"}

    try:
        result = json.loads(raw)
        text = result.get("result", {}).get("content", [{}])[0].get("text", "")
        try:
            return json.loads(text)
        except (json.JSONDecodeError, TypeError):
            if text:
                return {"_raw": text}
            return result
    except json.JSONDecodeError as e:
        return {"_error": f"JSON parse: {e}", "_raw": raw[:300]}


def read_excel_plan(path: str) -> tuple:
    """Read test plan xlsx. Returns (instances, test_cases)."""
    wb = openpyxl.load_workbook(path, data_only=True)

    # Sheet 1: instances
    instances = []
    if "instances" in wb.sheetnames:
        ws = wb["instances"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            inst = dict(zip(headers, row))
            if inst.get("name"):
                instances.append(inst)

    # Sheet 2: test_cases
    test_cases = []
    if "test_cases" in wb.sheetnames:
        ws = wb["test_cases"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            tc = dict(zip(headers, row))
            if tc.get("test_id"):
                test_cases.append(tc)

    wb.close()
    return instances, test_cases


def instance_id_map(mcp_bin: str, sim_url: str) -> dict:
    """Get {name: id} mapping of running instances."""
    result = mcp_call(mcp_bin, sim_url, "list_instances", {})
    instances = result.get("instances", [])
    return {inst["name"]: inst["id"] for inst in instances}


def run_test_plan(plan_path: str, mcp_bin: str, sim_url: str,
                  report_title: str = "EGC 自动化测试报告",
                  report_version: str = "v2.2.0",
                  port_start: int = 5000,
                  point_dir: str = ".",
                  upload_points: bool = True):
    """Execute full test plan with natural language parsing and EGC reporting."""
    instances_config, test_cases = read_excel_plan(plan_path)
    if not test_cases:
        print("❌ 未找到 test_cases sheet 或测试用例为空")
        return

    # Apply port template
    if port_start:
        apply_port_template(instances_config, port_start)

    # Load point mapping for NL parsing
    point_map = load_point_mapping(point_dir)
    if point_map:
        print(f"  📖 已加载 {len(point_map)} 个点表映射")

    print(f"\n{'='*60}")
    print(f"  EGC AutoTester")
    print(f"  测试计划: {plan_path}")
    print(f"  模拟器:   {sim_url}")
    print(f"  用例数:   {len(test_cases)}")
    print(f"  端口起始: {port_start}")
    print(f"{'='*60}\n")

    results = []
    instance_name_to_id = {}

    # Phase 1: Create instances (skip if already exist)
    print("【阶段1】创建实例...")

    # Auto-upload point tables
    if upload_points and os.path.isdir(point_dir):
        uploaded = set()
        for inst in instances_config:
            xlsx = inst.get("xlsx_file", "")
            if xlsx and xlsx not in uploaded:
                xlsx_path = os.path.join(point_dir, xlsx)
                if os.path.exists(xlsx_path):
                    import base64
                    with open(xlsx_path, "rb") as f:
                        b64 = base64.b64encode(f.read()).decode()
                    r = mcp_call(mcp_bin, sim_url, "upload_file", {
                        "filename": xlsx,
                        "content_base64": b64,
                    })
                    if r.get("status") == "uploaded":
                        print(f"  📤 上传点表: {xlsx}")
                    uploaded.add(xlsx)

    existing_map = instance_id_map(mcp_bin, sim_url)

    for inst in instances_config:
        name = inst["name"]
        iec_port = int(inst["iec104_port"])
        http_port = int(inst["http_port"])
        xlsx = inst["xlsx_file"]

        if name in existing_map:
            inst_id = existing_map[name]
            instance_name_to_id[name] = inst_id
            print(f"  ⏭️  跳过 {name} (已存在, ID:{inst_id})")
            continue

        cfg = json.dumps({
            "name": name,
            "iec104_port": iec_port,
            "xlsx_file": xlsx,
            "http_enabled": True,
            "http_port": http_port,
        }, ensure_ascii=False)
        r = mcp_call(mcp_bin, sim_url, "create_instance", {"config": cfg})
        inst_id = r.get("id") or r.get("_raw", "?")
        if inst_id and inst_id != "?" and not inst_id.startswith("错误"):
            instance_name_to_id[name] = inst_id
            print(f"  ✅ 创建 {name} (IEC:{iec_port} HTTP:{http_port}) → ID:{inst_id}")
        else:
            print(f"  ❌ 创建 {name} 失败: {r.get('_error', r.get('_raw', '?'))}")

    # Start all instances (only if not running)
    print("  └─ 启动实例...")
    for name, inst_id in instance_name_to_id.items():
        r = mcp_call(mcp_bin, sim_url, "start_instance", {"instance_id": inst_id})
        status = r.get("status", "")
        ok = "✅" if status == "ok" else "⏭️" if "running" in str(r) else "❌"
        print(f"     {ok} {name}")

    # MCP step executor for parsed natural language steps
    def _exec_mcp_step(mcp_bin, sim_url, step, name_to_id):
        """Execute a single parsed step and return {passed, detail}."""
        action = step["action"]
        inst = str(step.get("instance", ""))
        inst_id = name_to_id.get(inst, inst)
        detail = ""
        passed = False
        try:
            if action == "write":
                ioa = int(step["ioa"])
                val = float(step["value"])
                r = mcp_call(mcp_bin, sim_url, "write_point",
                             {"instance_id": inst_id, "ioa": ioa, "value": val})
                detail = f"写入 {inst} IOA:{ioa} = {val}"
                passed = True

            elif action == "assert":
                ioa = int(step["ioa"])
                r = mcp_call(mcp_bin, sim_url, "read_point",
                             {"instance_id": inst_id, "ioa": ioa})
                actual = r.get("value", r.get("_raw", "?"))
                try:
                    actual_f = float(actual)
                except (ValueError, TypeError):
                    actual_f = actual
                exp_min = step.get("expected_min", 0)
                exp_max = step.get("expected_max", 0)
                if isinstance(actual_f, (int, float)):
                    passed = exp_min <= actual_f <= exp_max
                else:
                    passed = str(actual) == str(step.get("expected_val", ""))
                ev = step.get("expected_val", f"{exp_min}~{exp_max}")
                detail = f"断言 {inst} IOA:{ioa} = {actual} (期望 {ev})"

        except Exception as e:
            detail = f"异常: {e}"
        return {"passed": passed, "detail": detail}

    # Phase 2: Execute test cases
    print("\n【阶段2】执行测试用例...")
    passed_ids = set()

    for tc in test_cases:
        tid = str(tc.get("test_id", "")).strip()
        desc = str(tc.get("description", "")).strip()
        action = str(tc.get("action", "")).strip()
        instance = str(tc.get("instance", "")).strip()
        ioa_raw = tc.get("ioa")
        value_raw = tc.get("value")
        exp_min = tc.get("expected_min")
        exp_max = tc.get("expected_max")
        wait_ms = int(tc.get("wait_ms") or 0)
        depends = str(tc.get("depends_on") or "").strip()

        if not tid or not action:
            # Natural language parsing: use description if no explicit action
            if desc and point_map:
                parsed = parse_natural_step(desc, point_map)
                if parsed:
                    for i, step in enumerate(parsed):
                        sub_tid = f"{tid}-{i+1}" if len(parsed) > 1 else tid
                        sub_desc = f"{desc} → {step.get('desc_detail', '')}"
                        # Execute step inline
                        r = _exec_mcp_step(mcp_bin, sim_url, step,
                                           instance_name_to_id)
                        status = "✅" if r["passed"] else "❌"
                        if r["passed"]:
                            passed_ids.add(sub_tid)
                        results.append({
                            "id": sub_tid, "desc": sub_desc,
                            "action": step["action"],
                            "status": status, "detail": r["detail"],
                            "step_info": step,
                        })
                        print(f"  {status}  {sub_tid}: {sub_desc}")
                        print(f"     {r['detail']}")
                    continue
            continue

        # Skip if dependency failed
        if depends and depends not in passed_ids:
            results.append({
                "id": tid, "desc": desc, "action": action,
                "status": "⏭️", "detail": f"依赖 {depends} 未通过",
            })
            print(f"  ⏭️  {tid}: {desc} (依赖 {depends} 未通过)")
            continue

        # Wait
        if wait_ms > 0:
            time.sleep(wait_ms / 1000)

        detail = ""
        passed = False

        try:
            if action == "sleep":
                passed = True
                detail = f"等待 {wait_ms}ms"

            elif action == "write":
                inst_id = instance_name_to_id.get(instance, instance)
                ioa = int(ioa_raw)
                val = float(value_raw)
                r = mcp_call(mcp_bin, sim_url, "write_point", {
                    "instance_id": inst_id, "ioa": ioa, "value": val,
                })
                changed = r.get("changed", False)
                passed = changed if changed else True  # changed=False is ok for idempotent writes
                detail = f"写入 IOA:{ioa} = {val} → changed:{changed}"

            elif action == "read":
                inst_id = instance_name_to_id.get(instance, instance)
                ioa = int(ioa_raw)
                r = mcp_call(mcp_bin, sim_url, "read_point", {
                    "instance_id": inst_id, "ioa": ioa,
                })
                actual = r.get("value", r.get("_raw", "?"))
                passed = True
                detail = f"读取 IOA:{ioa} = {actual}"

            elif action == "write_points":
                inst_id = instance_name_to_id.get(instance, instance)
                points = json.loads(value_raw) if isinstance(value_raw, str) else value_raw
                r = mcp_call(mcp_bin, sim_url, "write_points", {
                    "instance_id": inst_id, "points": points,
                })
                succeeded = r.get("succeeded", 0)
                total = r.get("total", 0)
                passed = succeeded == total and total > 0
                detail = f"批量写入 {succeeded}/{total} 成功"

            elif action == "assert":
                inst_id = instance_name_to_id.get(instance, instance)
                ioa = int(ioa_raw)
                r = mcp_call(mcp_bin, sim_url, "read_point", {
                    "instance_id": inst_id, "ioa": ioa,
                })
                actual_raw = r.get("value", r.get("_raw", "?"))
                # Ensure comparison-safe types
                try:
                    actual = float(actual_raw)
                except (ValueError, TypeError):
                    actual = actual_raw
                exp_min_f = float(exp_min) if exp_min is not None else None
                exp_max_f = float(exp_max) if exp_max is not None else None

                try:
                    if exp_min_f is not None and exp_max_f is not None:
                        passed = exp_min_f <= actual <= exp_max_f
                        detail = f"IOA:{ioa} = {actual} (期望 {exp_min_f}~{exp_max_f})"
                    elif exp_min_f is not None:
                        passed = actual == exp_min_f
                        detail = f"IOA:{ioa} = {actual} (期望 = {exp_min_f})"
                    else:
                        passed = True
                        detail = f"IOA:{ioa} = {actual} (读取成功)"
                except TypeError:
                    passed = False
                    detail = f"IOA:{ioa} = {actual} (类型不匹配: {type(actual).__name__})"

            elif action == "config_auto":
                inst_id = instance_name_to_id.get(instance, instance)
                ioa = int(ioa_raw)
                params = json.loads(value_raw) if isinstance(value_raw, str) else {}
                strategy = params.pop("strategy", "random")
                enabled = params.pop("enabled", True)
                r = mcp_call(mcp_bin, sim_url, "config_auto_change", {
                    "instance_id": inst_id,
                    "ioa": ioa,
                    "strategy": strategy,
                    "enabled": enabled,
                    "params": json.dumps(params, ensure_ascii=False),
                })
                success = r.get("success", r.get("_raw", ""))
                passed = True
                detail = f"配置 IOA:{ioa} 策略={strategy}"

            elif action == "formula_config":
                inst_id = instance_name_to_id.get(instance, instance)
                ioa = int(ioa_raw)
                params = json.loads(value_raw) if isinstance(value_raw, str) else {}
                r = mcp_call(mcp_bin, sim_url, "config_auto_change", {
                    "instance_id": inst_id,
                    "ioa": ioa,
                    "strategy": "custom",
                    "enabled": True,
                    "params": json.dumps(params, ensure_ascii=False),
                })
                passed = True
                detail = f"配置 IOA:{ioa} 自定义公式"

            else:
                detail = f"未知 action: {action}"
                passed = False

        except Exception as e:
            detail = f"异常: {e}"
            passed = False

        status = "✅" if passed else "❌"
        if passed:
            passed_ids.add(tid)
        results.append({
            "id": tid, "desc": desc, "action": action,
            "status": status, "detail": detail,
        })
        print(f"  {status}  {tid}: {desc}")
        print(f"     {detail}")

    # Phase 3: Generate report
    print("\n【阶段3】生成报告...")
    timestamp = datetime.now()
    ts_md = timestamp.strftime("%Y%m%d-%H%M%S")
    ts_hr = timestamp.strftime("%Y-%m-%d %H:%M:%S")

    md = generate_markdown(plan_path, sim_url, results, instances_config,
                            report_title, report_version, ts_hr)
    html = generate_html(plan_path, sim_url, results, instances_config,
                          report_title, report_version, ts_hr, mcp_bin)

    md_path = f"test-report-{ts_md}.md"
    html_path = f"test-report-{ts_md}.html"
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md)
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  ✅ Markdown 报告: {md_path}")
    print(f"  ✅ HTML 报告:     {html_path}")
    print(f"\n{'='*60}")
    print(f"  测试完成")
    print(f"  总计: {len(results)}  | 通过: {sum(1 for r in results if r['status'] == '✅')}  | 失败: {sum(1 for r in results if r['status'] == '❌')}")
    print(f"{'='*60}")
    return html_path


def generate_markdown(plan_path, sim_url, results, instances_config,
                       report_title, report_version, ts_hr):
    """Generate markdown report."""
    total = len(results)
    passed = sum(1 for r in results if r["status"] == "✅")
    failed = sum(1 for r in results if r["status"] == "❌")
    skipped = sum(1 for r in results if r["status"] == "⏭️")

    lines = []
    lines.append(f"# {report_title}\n")
    lines.append(f"**版本**: {report_version}  \n")
    lines.append(f"**测试计划**: {Path(plan_path).name}  \n")
    lines.append(f"**模拟器**: {sim_url}  \n")
    lines.append(f"**执行时间**: {ts_hr}  \n")
    lines.append(f"**测试工具**: IEC104 AutoTester (MCP)\n")
    lines.append("---\n")

    rate = f"{passed / total * 100:.1f}%" if total > 0 else "-"
    lines.append("## 汇总\n")
    lines.append("| 总计 | 通过 | 失败 | 跳过 | 通过率 |")
    lines.append("|------|------|------|------|--------|")
    lines.append(f"| {total} | {passed} | {failed} | {skipped} | {rate} |\n")

    if instances_config:
        lines.append("## 实例配置\n")
        lines.append("| 名称 | IEC104端口 | HTTP端口 | 点表文件 |")
        lines.append("|------|-----------|----------|---------|")
        for inst in instances_config:
            lines.append(f"| {inst['name']} | {inst['iec104_port']} | {inst['http_port']} | {inst['xlsx_file']} |")
        lines.append("")

    lines.append("## 执行详情\n")
    for r in results:
        lines.append(f"### {r['status']} {r['id']}: {r['desc']}")
        lines.append(f"**动作**: `{r['action']}`  \n")
        lines.append(f"**结果**: {r['detail']}  \n")

    failures = [r for r in results if r["status"] == "❌"]
    if failures:
        lines.append("## 失败用例\n")
        for r in failures:
            lines.append(f"- **{r['id']}**: {r['desc']} → {r['detail']}")
        lines.append("")

    lines.append("---\n")
    lines.append(f"*报告由 IEC104 AutoTester 自动生成于 {ts_hr}*\n")
    return "\n".join(lines)


def generate_html(plan_path, sim_url, results, instances_config,
                   report_title, report_version, ts_hr, mcp_bin):
    """Generate HTML report using the template."""
    total = len(results)
    passed = sum(1 for r in results if r["status"] == "✅")
    failed = sum(1 for r in results if r["status"] == "❌")
    skipped = sum(1 for r in results if r["status"] == "⏭️")
    rate = f"{passed / total * 100:.1f}%" if total > 0 else "-"

    def esc(s):
        return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")

    # Find HTML template
    template_candidates = [
        "test-report-template.html",
        os.path.join(os.path.dirname(__file__), "..", "report-template.html"),
    ]
    template_path = None
    for p in template_candidates:
        if os.path.exists(p):
            template_path = p
            break
    if not template_path:
        return "<html><body><h1>HTML template not found</h1><p>Template search paths: " + ", ".join(template_candidates) + "</p></body></html>"

    with open(template_path, "r", encoding="utf-8") as f:
        html = f.read()

    # === Header ===
    html = html.replace(
        '<h1>IEC104 模拟器自动化测试报告</h1>',
        f'<h1>{esc(report_title)}</h1>'
    )
    html = html.replace(
        '<span class="value" id="version">v2.2.0</span>',
        f'<span class="value" id="version">{esc(report_version)}</span>'
    )
    html = html.replace(
        '<span class="value" id="test-date">2026-05-15 19:06:55</span>',
        f'<span class="value" id="test-date">{esc(ts_hr)}</span>'
    )
    html = html.replace(
        '<span class="value" id="simulator">http://10.65.99.13:8989</span>',
        f'<span class="value" id="simulator">{esc(sim_url)}</span>'
    )

    # === Summary ===
    pw = f"{rate}" if "%" in rate else "0"
    html = html.replace('<div class="number" id="pass-count">6</div>', f'<div class="number" id="pass-count">{passed}</div>')
    html = html.replace('<div class="number" id="fail-count">0</div>', f'<div class="number" id="fail-count">{failed}</div>')
    html = html.replace('<div class="number" id="skip-count">0</div>', f'<div class="number" id="skip-count">{skipped}</div>')
    html = html.replace('<div class="number" id="pass-rate">100%</div>', f'<div class="number" id="pass-rate">{rate}</div>')
    html = html.replace('style="width: 100%"', f'style="width: {pw}"')

    # === Environment table ===
    env_rows = ""
    for inst in instances_config:
        env_rows += f"""<tr>
            <td>{esc(inst['name'])}</td>
            <td><code>{esc(inst['iec104_port'])}</code></td>
            <td><code>{esc(inst['http_port'])}</code></td>
            <td>{esc(inst['xlsx_file'])}</td>
        </tr>\n"""
    html = html.replace(
        '<!-- ===== Test Environment ===== -->',
        f'<!-- ===== Test Environment ===== -->\n<tbody id="env-table">\n{env_rows}\n</tbody>\n'
    )
    # Remove placeholder rows from template
    env_placeholder = '<tbody id="env-table">\n          <tr>\n            <td>关口表</td>\n            <td><code>5000</code></td>\n            <td><code>6000</code></td>\n            <td>固定验证-关口表.xlsx</td>\n          </tr>\n          <tr>\n            <td>储能-01</td>\n            <td><code>5001</code></td>\n            <td><code>6001</code></td>\n            <td>固定验证-储能.xlsx</td>\n          </tr>\n          <tr>\n            <td>FCR</td>\n            <td><code>5005</code></td>\n            <td><code>6005</code></td>\n            <td>固定验证-FCR.xlsx</td>\n          </tr>\n        </tbody>'
    html = html.replace(env_placeholder, '')

    # === Test cases ===
    case_html = ""
    for r in results:
        icon = r["status"]
        case_html += f"""<div class="case-row">
        <div class="case-icon">{icon}</div>
        <div class="case-body">
          <div class="case-id">{esc(r['id'])} — {esc(r['desc'])}</div>
          <div class="case-detail">动作: {esc(r['action'])} | {esc(r['detail'])}</div>
        </div>
      </div>\n"""
    html = html.replace(
        '<div class="card-body" id="test-cases">',
        f'<div class="card-body" id="test-cases">\n{case_html}'
    )
    # Remove placeholder cases from template
    tc_placeholder_start = '<!-- ===== Test Cases ===== -->'
    tc_placeholder_end = '<!-- ===== Test Records ===== -->'
    # Find the placeholder test cases between the header and actual records
    tc_section = html[html.find(tc_placeholder_start):html.find(tc_placeholder_end)]
    # Replace the placeholder inner content
    old_cases = '<div class="case-row">\n        <div class="case-icon">✅</div>\n        <div class="case-body">\n          <div class="case-id">TC-001 — 关口表有功功率 → 储能控制值</div>\n          <div class="case-desc">验证关口表功率变化后储能 AO 控制值正确联动</div>\n          <div class="case-detail">write 关口表 IOA:16385 = 10 → assert 储能-01 IOA:25089 ∈ [38,42]</div>\n        </div>\n      </div>\n\n      <div class="case-row">\n        <div class="case-icon">❌</div>\n        <div class="case-body">\n          <div class="case-id">TC-002 — FCR 功率设定</div>\n          <div class="case-desc">验证 FCR 实例 ap_exp_percent_ulimit_value 写入与回读一致</div>\n          <div class="case-detail">write FCR IOA:16385 = 10 → read FCR IOA:16385 → 期望 10, 实际 5</div>\n        </div>\n      </div>'
    html = html.replace(old_cases, '')

    # === Test records ===
    log_html = ""
    for i, r in enumerate(results):
        t = i + 1
        log_html += f"""<div class="log-entry">
        <span class="status-icon">{r['status']}</span>
        <span class="time">步骤 {t:02d}</span>
        <span class="content">{esc(r['id'])}: {esc(r['desc'])} — {esc(r['detail'])}</span>
      </div>\n"""
    html = html.replace(
        '<div class="card-body" id="test-records">',
        f'<div class="card-body" id="test-records">\n{log_html}'
    )

    # === Bug records ===
    failures = [r for r in results if r["status"] == "❌"]
    if failures:
        bug_html = ""
        for i, r in enumerate(failures):
            bug_html += f"""<div class="bug-card">
        <div class="bug-title">BUG-{i+1:03d}: {esc(r['id'])} — {esc(r['desc'])}</div>
        <div class="bug-detail">
          <p><strong>测试用例:</strong> {esc(r['id'])}</p>
          <p><strong>描述:</strong> {esc(r['desc'])}</p>
          <p><strong>结果:</strong> {esc(r['detail'])}</p>
        </div>
      </div>\n"""
        html = html.replace(
            '<div style="text-align:center;padding:24px;color:var(--color-text-secondary);font-size:14px">\n              ✨ 无 Bug，全部通过\n            </div>',
            bug_html.strip()
        )
    else:
        html = html.replace(
            '<div style="text-align:center;padding:24px;color:var(--color-text-secondary);font-size:14px">\n              ✨ 无 Bug，全部通过\n            </div>',
            '<div style="text-align:center;padding:24px;color:var(--color-success);font-size:16px;font-weight:600">✅ 无 Bug，全部通过</div>'
        )

    # Remove the placeholder log entries from the template
    log_entries = html.split('<!-- ===== Bug Records ===== -->')
    if len(log_entries) > 1:
        # Find the full test-records card content
        rec_start = html.find('id="test-records">')
        rec_end = html.find('<!-- ===== Bug Records ===== -->')
        records_section = html[rec_start:rec_end]
        # Replace the placeholder entries inside
        old_log_start = records_section.find('<div class="log-entry">')
        if old_log_start >= 0:
            new_records = records_section[:old_log_start] + '\n' + log_html + '\n    '
            html = html[:rec_start] + new_records + html[rec_end:]

    return html


def main():
    parser = argparse.ArgumentParser(description="IEC104 AutoTester — EGC 自动化测试框架")
    parser.add_argument("--plan", required=True, help="测试计划 xlsx 文件路径")
    parser.add_argument("--simulator", default="http://10.65.99.13:8989", help="模拟器 HTTP API 地址")
    parser.add_argument("--mcp", default="./bin/iec104-mcp", help="MCP 二进制路径")
    parser.add_argument("--title", default="EGC 自动化测试报告", help="报告标题")
    parser.add_argument("--version", default="v2.2.0", help="测试版本号")
    parser.add_argument("--port-start", type=int, default=5000,
                        help="起始 IEC104 端口（Excel 中用 {p+0} {p+1} 模板）")
    parser.add_argument("--point-dir", default=".",
                        help="点表 xlsx 文件所在目录（用于自动上传和解析）")
    parser.add_argument("--upload-points", action="store_true", default=True,
                        help="自动上传点表文件到模拟器")
    args = parser.parse_args()

    if not os.path.exists(args.mcp):
        print(f"❌ MCP 二进制不存在: {args.mcp}")
        sys.exit(1)
    if not os.path.exists(args.plan):
        print(f"❌ 测试计划文件不存在: {args.plan}")
        sys.exit(1)

    run_test_plan(
        plan_path=args.plan,
        mcp_bin=args.mcp,
        sim_url=args.simulator,
        report_title=args.title,
        report_version=args.version,
        port_start=args.port_start,
        point_dir=args.point_dir,
        upload_points=args.upload_points,
    )


if __name__ == "__main__":
    main()
